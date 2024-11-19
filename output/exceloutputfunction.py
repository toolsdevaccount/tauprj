from decimal import Decimal
from django.shortcuts import redirect
#Excel
import openpyxl
#from django.conf.urls.static import static
from django.conf import settings
from decimal import Decimal
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
# メッセージ
from django.contrib import messages
#LOG出力設定
import logging
logger = logging.getLogger(__name__)

def exceloutput(request, table, fileNameDate, ManagerCode):
    # Excelのテンプレートファイルの読み込み
    wb = openpyxl.load_workbook(settings.MEDIA_ROOT + '/Contract.xlsx') 

    #黒い実線
    side = Side(style='thin', color='000000')
    #周囲に線を引く
    border_aro = Border(top=side, bottom=side, left=side, right=side)
    #背景色グレー
    fill_gray = PatternFill(patternType='solid', fgColor='d3d3d3', bgColor='d3d3d3')

    i=2
    #小計出現カウント
    j=0
    row=2
    rows=[]
    comparison =''
    for book in table:
        sheet = wb['Sheet1']
        font = Font(name='メイリオ')
        sheet.freeze_panes = 'A2'
        if book['OrderingTableId__SlipDiv'] != comparison and i!=2 and j!=0:
            cell = sheet['A' + str(i) ]
            cell.font = font
            cell.value = comparison + str('-小計') 
            cell.border = border_aro
            cell.fill = fill_gray

            cell = sheet['B' + str(i) ]
            cell.font = font
            cell.value = '' 
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['C' + str(i) ]
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['D' + str(i) ] 
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['E' + str(i) ]
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['F' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = '=SUM(F' + str(row) + ':F'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['G' + str(i) ]
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['H' + str(i) ]
            cell.font = font           
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(H' + str(row) + ':H'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['I' + str(i) ]
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['J' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(J' + str(row) + ':J'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['K' + str(i) ]
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['L' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(L' + str(row) + ':L'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['M' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(M' + str(row) + ':M'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['N' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = '=SUM(N' + str(row) + ':N'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['O' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(O' + str(row) + ':O'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['P' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = '=SUM(P' + str(row) + ':P'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['Q' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(Q' + str(row) + ':Q'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['R' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = '=SUM(R' + str(row) + ':R'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['S' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(S' + str(row) + ':S'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['T' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = '=SUM(T' + str(row) + ':T'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['U' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(U' + str(row) + ':U'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['V' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = '=SUM(V' + str(row) + ':V'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['W' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=SUM(W' + str(row) + ':W'+ str(i-1) + ')'
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['X' + str(i) ]
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            cell = sheet['Y' + str(i) ]
            cell.font = font
            cell.value = ''
            cell.fill = fill_gray
            cell.border = border_aro

            rows.append(str(i))
            i+=1
            j=0
            row=str(i)
        if book['DetailVolume'] != Decimal('0.00'):
            cell = sheet['A' + str(i) ]
            cell.font = font
            cell.value = book['OrderingTableId__SlipDiv'] + '-' + book['OrderingTableId__OrderNumber'] 
            cell.border = border_aro

            cell = sheet['B' + str(i) ]
            cell.font = font
            cell.value = book['RequestCustomer'] 
            cell.border = border_aro

            cell = sheet['C' + str(i) ]
            cell.font = font
            cell.value = book['Manager_firstname'] + ' ' + book['Manager_lastname']
            cell.border = border_aro

            cell = sheet['D' + str(i) ] 
            cell.font = font
            cell.value = book['ProductName']
            cell.border = border_aro

            cell = sheet['E' + str(i) ]
            cell.font = font
            cell.value = book['OrderingCount']
            cell.border = border_aro

            cell = sheet['F' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = book['DetailVolume']
            cell.border = border_aro

            cell = sheet['G' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = book['DetailUnitPrice']
            cell.border = border_aro

            cell = sheet['H' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(F' + str(i) + '*G' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['I' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = book['ProcessingUnitPrice']
            cell.border = border_aro

            cell = sheet['J' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(F' + str(i) + '*I' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['K' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = book['DetailSellPrice']
            cell.border = border_aro

            cell = sheet['L' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(F' + str(i) + '*K' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['M' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=L' + str(i) + '-H' + str(i) + '-J' + str(i)
            cell.border = border_aro

            cell = sheet['N' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = 0
            cell.border = border_aro

            cell = sheet['O' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(K' + str(i) + '*N' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['P' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = 0
            cell.border = border_aro

            cell = sheet['Q' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(K' + str(i) + '*P' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['R' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = 0
            cell.border = border_aro

            cell = sheet['S' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(K' + str(i) + '*R' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['T' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = 0
            cell.border = border_aro

            cell = sheet['U' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(K' + str(i) + '*T' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['V' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0.00;[red]-#,##0.00'
            cell.value = 0
            cell.border = border_aro

            cell = sheet['W' + str(i) ]
            cell.font = font
            cell.number_format = '#,##0;[red]-#,##0'
            cell.value = '=ROUNDDOWN(K' + str(i) + '*V' + str(i) + ',0)'
            cell.border = border_aro

            cell = sheet['X' + str(i) ]
            cell.font = font
            cell.number_format = 'yyyy年mm月dd日'
            cell.value = book['SpecifyDeliveryDate']
            cell.border = border_aro

            cell = sheet['Y' + str(i) ]
            cell.font = font
            cell.number_format = 'yyyy年mm月dd日'
            cell.value = book['StainAnswerDeadline']
            cell.border = border_aro

            i+=1
            j+=1

        comparison = book['OrderingTableId__SlipDiv']

    #最後の小計
    if j!=0:
        cell = sheet['A' + str(i) ]
        cell.font = font
        cell.value = comparison + str('-小計') 
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['B' + str(i) ]
        cell.font = font
        cell.value = '' 
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['C' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['D' + str(i) ] 
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['E' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['F' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = '=SUM(F' + str(row) + ':F'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['G' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['H' + str(i) ]
        cell.font = font           
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(H' + str(row) + ':H'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['I' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['J' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(J' + str(row) + ':J'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['K' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['L' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(L' + str(row) + ':L'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['M' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(M' + str(row) + ':M'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['N' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = '=SUM(N' + str(row) + ':N'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['O' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(O' + str(row) + ':O'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['P' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = '=SUM(P' + str(row) + ':P'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['Q' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(Q' + str(row) + ':Q'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['R' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = '=SUM(R' + str(row) + ':R'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['S' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(S' + str(row) + ':S'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['T' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = '=SUM(T' + str(row) + ':T'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['U' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(U' + str(row) + ':U'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['V' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = '=SUM(V' + str(row) + ':V'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['W' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=SUM(W' + str(row) + ':W'+ str(i-1) + ')'
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['X' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

        cell = sheet['Y' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.fill = fill_gray
        cell.border = border_aro

    #総合計
    rows.append(str(i))
    cell = sheet['A' + str(i+1) ]
    cell.font = font
    cell.value = str('総合計') 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['B' + str(i+1) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['C' + str(i+1) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['D' + str(i+1) ] 
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['E' + str(i+1) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['F' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=F' + str(rows[t])
        else:
            DetailVolume+= '+F' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['G' + str(i+1) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['H' + str(i+1) ]
    cell.font = font           
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=H' + str(rows[t])
        else:
            DetailVolume+= '+H' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['I' + str(i+1) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['J' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=J' + str(rows[t])
        else:
            DetailVolume+= '+J' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['K' + str(i+1) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['L' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=L' + str(rows[t])
        else:
            DetailVolume+= '+L' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['M' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=M' + str(rows[t])
        else:
            DetailVolume+= '+M' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['N' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=N' + str(rows[t])
        else:
            DetailVolume+= '+N' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['O' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=O' + str(rows[t])
        else:
            DetailVolume+= '+O' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['P' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=P' + str(rows[t])
        else:
            DetailVolume+= '+P' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['Q' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=Q' + str(rows[t])
        else:
            DetailVolume+= '+Q' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['R' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=R' + str(rows[t])
        else:
            DetailVolume+= '+R' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['S' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=S' + str(rows[t])
        else:
            DetailVolume+= '+S' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['T' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=T' + str(rows[t])
        else:
            DetailVolume+= '+T' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['U' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=U' + str(rows[t])
        else:
            DetailVolume+= '+U' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['V' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=V' + str(rows[t])
        else:
            DetailVolume+= '+V' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['W' + str(i+1) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    for t in range(len(rows)):
        if t==0:
            DetailVolume= '=W' + str(rows[t])
        else:
            DetailVolume+= '+W' + str(rows[t])
    cell.value = DetailVolume
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['X' + str(i+1) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['Y' + str(i+1) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    # Excelを返すためにcontent_typeに「application/vnd.ms-excel」をセットします。
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' %  fileNameDate[1] + '-' + str(ManagerCode) + '.xlsx'

    # データの書き込みを行なったExcelファイルを保存する
    wb.save(response)

    return response
