from decimal import Decimal
from django.shortcuts import redirect
#Excel
import openpyxl
#from django.conf.urls.static import static
from django.conf import settings
from decimal import Decimal
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
# メッセージ
from django.contrib import messages
#LOG出力設定
import logging
logger = logging.getLogger(__name__)

def exceloutput(request, table, fileNameDate):
    # Excelのテンプレートファイルの読み込み
    wb = openpyxl.load_workbook(settings.MEDIA_ROOT + '/StockList.xlsx') 

    #黒い実線
    side = Side(style='thin', color='000000')
    #周囲に線を引く
    border_aro = Border(top=side, bottom=side, left=side, right=side)
    #背景色グレー
    fill_gray = PatternFill(patternType='solid', fgColor='d3d3d3', bgColor='d3d3d3')

    i=2
    row=2
    rows=[]
    for book in table:
        sheet = wb['Sheet1']
        font = Font(name='メイリオ',size=9)
        sheet.freeze_panes = 'F2'

        cell = sheet['A' + str(i) ]
        cell.font = font
        cell.value = book['OrderingId__OrderNumber'] 
        cell.border = border_aro

        cell = sheet['B' + str(i) ]
        cell.font = font
        cell.value = book['ProductName'] 
        cell.border = border_aro

        cell = sheet['C' + str(i) ]
        cell.font = font
        cell.value = book['ResultItemNumber'] 
        cell.border = border_aro

        cell = sheet['D' + str(i) ] 
        cell.font = font
        cell.value = book['OrderingCount']
        cell.border = border_aro

        cell = sheet['E' + str(i) ]
        cell.font = font
        cell.value = book['DetailColor']
        cell.border = border_aro

        cell = sheet['F' + str(i) ]
        cell.font = font
        cell.value = book['Manager_firstname'] + ' ' + book['Manager_lastname']
        cell.border = border_aro

        cell = sheet['G' + str(i) ]
        cell.font = font
        cell.value = book['RequestCustomerCode']
        cell.border = border_aro

        cell = sheet['H' + str(i) ]
        cell.font = font
        cell.value = book['RequestCustomer']
        cell.border = border_aro

        cell = sheet['I' + str(i) ]
        cell.font = font
        cell.value = book['ShippingCustomerCode']
        cell.border = border_aro

        cell = sheet['J' + str(i) ]
        cell.font = font
        cell.value = book['ShippingCustomer']
        cell.border = border_aro

        cell = sheet['K' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = book['CarryForward_total']
        cell.border = border_aro

        cell = sheet['L' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = book['ReciveStock']
        cell.border = border_aro

        cell = sheet['M' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = book['Issue']
        cell.border = border_aro

        cell = sheet['N' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = book['Remaining']
        cell.border = border_aro

        cell = sheet['O' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = book['DetailUnitPrice']
        cell.border = border_aro

        cell = sheet['P' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=ROUNDDOWN(O' + str(i) + '*N' + str(i) + ',0)'
        cell.border = border_aro

        cell = sheet['Q' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = book['Process_total']
        cell.border = border_aro

        cell = sheet['R' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0.00;[red]-#,##0.00'
        cell.value = book['Process']
        cell.border = border_aro

        cell = sheet['S' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = book['ProcessingUnitprice']
        cell.border = border_aro

        cell = sheet['T' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=ROUNDDOWN(R' + str(i) + '*S' + str(i) + ',0)'
        cell.border = border_aro

        cell = sheet['U' + str(i) ]
        cell.font = font
        cell.number_format = '#,##0;[red]-#,##0'
        cell.value = '=ROUNDDOWN(P' + str(i) + '+T' + str(i) + ',0)'
        cell.border = border_aro

        cell = sheet['V' + str(i) ]
        cell.font = font
        cell.value = ''
        cell.border = border_aro

        i+=1

    #総合計
    rows.append(str(i))
    cell = sheet['A' + str(i) ]
    cell.font = font
    cell.value = str('総合計') 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['B' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['C' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['D' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['E' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['F' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['G' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['H' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['I' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    rows.append(str(i))
    cell = sheet['J' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['K' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    cell.value = '=SUBTOTAL(109,K' + str(row) + ':K'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['L' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    cell.value = '=SUBTOTAL(109,L' + str(row) + ':L'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['M' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    cell.value = '=SUBTOTAL(109,M' + str(row) + ':M'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['N' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    cell.value = '=SUBTOTAL(109,N' + str(row) + ':N'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['O' + str(i) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['P' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    cell.value = '=SUBTOTAL(109,P' + str(row) + ':P'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['Q' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    cell.value = '=SUBTOTAL(109,Q' + str(row) + ':Q'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['R' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0.00;[red]-#,##0.00'
    cell.value = '=SUBTOTAL(109,R' + str(row) + ':R'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['S' + str(i) ]
    cell.font = font
    cell.value = ''
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['T' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    cell.value = '=SUBTOTAL(109,T' + str(row) + ':T'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['U' + str(i) ]
    cell.font = font
    cell.number_format = '#,##0;[red]-#,##0'
    cell.value = '=SUBTOTAL(109,U' + str(row) + ':U'+ str(i-1) + ')'
    cell.fill = fill_gray
    cell.border = border_aro

    cell = sheet['V' + str(i) ]
    cell.font = font
    cell.value = '' 
    cell.fill = fill_gray
    cell.border = border_aro

    #オートフィルター設定
    ws = wb['Sheet1']
    ws.auto_filter.ref = 'A1:V' + str(i)

    # Excelを返すためにcontent_typeに「application/vnd.ms-excel」をセットします。
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=%s' %  fileNameDate[1] + '_StockList' + '.xlsx'

    # データの書き込みを行なったExcelファイルを保存する
    wb.save(response)

    return response
